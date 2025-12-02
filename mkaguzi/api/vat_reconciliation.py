"""
VAT Reconciliation API
Handles VAT reconciliation between System Data, iTax Data, and TIMs Device Data
"""

import frappe
from frappe import _
import pandas as pd
import os
from datetime import datetime
import json
from io import BytesIO

# Amount matching tolerance (±0.02 to account for rounding differences between systems)
AMOUNT_TOLERANCE = 0.02

# Threshold for background job processing (lowered from 10000 to 5000 for better performance)
LARGE_FILE_THRESHOLD = 5000


# Expected CSV headers for each file type
EXPECTED_HEADERS = {
    "system": {
        "required": ["posting_date", "cu_invoice_number", "net_amount"],
        "optional": ["invoice_number"],
        "aliases": {
            "posting_date": ["posting_date", "date", "posting date", "transaction_date"],
            "invoice_number": ["invoice_number", "invoice_no", "invoice no", "inv_no"],
            "cu_invoice_number": ["cu_invoice_number", "cu_invoice_no", "cu invoice number", "cu_inv_no", "etims_invoice"],
            "net_amount": ["net_amount", "amount", "net amount", "value", "total"]
        }
    },
    "itax": {
        "required": ["posting_date", "cu_invoice_number", "amount"],
        "optional": [],
        "aliases": {
            "posting_date": ["posting_date", "date", "posting date", "transaction_date"],
            "cu_invoice_number": ["cu_invoice_number", "cu_invoice_no", "cu invoice number", "cu_inv_no", "etims_invoice"],
            "amount": ["amount", "net_amount", "net amount", "value", "total", "taxable_amount"]
        }
    },
    "tims": {
        "required": ["posting_date", "cu_invoice_number", "amount"],
        "optional": [],
        "aliases": {
            "posting_date": ["posting_date", "date", "posting date", "transaction_date"],
            "cu_invoice_number": ["cu_invoice_number", "cu_invoice_no", "cu invoice number", "cu_inv_no", "etims_invoice"],
            "amount": ["amount", "net_amount", "value", "total", "invoice_amount"]
        }
    }
}


@frappe.whitelist()
def get_vat_reconciliations(filters=None, page=1, page_size=50):
    """
    Get list of VAT reconciliations with pagination and filtering
    """
    try:
        filter_conditions = {}
        
        if filters:
            data = frappe.parse_json(filters) if isinstance(filters, str) else filters
            
            if data.get("reconciliation_month"):
                filter_conditions["reconciliation_month"] = data["reconciliation_month"]
            if data.get("fiscal_year"):
                filter_conditions["fiscal_year"] = data["fiscal_year"]
            if data.get("reconciliation_type"):
                filter_conditions["reconciliation_type"] = data["reconciliation_type"]
            if data.get("status"):
                filter_conditions["status"] = data["status"]
            if data.get("search"):
                filter_conditions["reconciliation_id"] = ["like", f"%{data['search']}%"]

        offset = (int(page) - 1) * int(page_size)

        reconciliations = frappe.get_all(
            "VAT Reconciliation",
            filters=filter_conditions,
            fields=[
                "name", "reconciliation_id", "reconciliation_type", "status",
                "reconciliation_month", "fiscal_year", "reconciliation_date",
                "total_source_a_records", "total_source_b_records",
                "total_source_a_amount", "total_source_b_amount",
                "total_matched", "total_unmatched_source_a", "total_unmatched_source_b",
                "total_amount_discrepancies", "total_variance_amount", "match_percentage",
                "reconciliation_status", "prepared_by", "is_background_job",
                "progress_percent", "completed_at", "creation", "modified",
                "system_data_file", "itax_data_file", "tims_data_file"
            ],
            order_by="modified desc",
            limit_page_length=int(page_size),
            limit_start=offset
        )

        total_count = frappe.db.count("VAT Reconciliation", filters=filter_conditions)

        return {
            "items": reconciliations,
            "total_count": total_count,
            "page": int(page),
            "page_size": int(page_size),
            "total_pages": (total_count + int(page_size) - 1) // int(page_size)
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "VAT Reconciliation List Error")
        frappe.throw(str(e))


# Mapping from frontend values to DocType select options
RECONCILIATION_TYPE_MAP = {
    "system_vs_itax": "System vs iTax",
    "system_vs_tims": "System vs TIMs",
    "itax_vs_tims": "iTax vs TIMs",
    # Also accept the proper values directly
    "System vs iTax": "System vs iTax",
    "System vs TIMs": "System vs TIMs",
    "iTax vs TIMs": "iTax vs TIMs"
}


@frappe.whitelist()
def create_reconciliation(reconciliation_type, reconciliation_month, fiscal_year):
    """
    Create a new VAT Reconciliation document
    """
    try:
        # Map the reconciliation type to proper format
        mapped_type = RECONCILIATION_TYPE_MAP.get(reconciliation_type)
        if not mapped_type:
            frappe.throw(frappe._("Invalid reconciliation type: {0}").format(reconciliation_type))
        
        doc = frappe.get_doc({
            "doctype": "VAT Reconciliation",
            "reconciliation_type": mapped_type,
            "reconciliation_month": reconciliation_month,
            "fiscal_year": fiscal_year,
            "status": "Draft",
            "reconciliation_date": frappe.utils.today()
        })
        doc.insert()
        frappe.db.commit()

        return {
            "success": True,
            "name": doc.name,
            "reconciliation_id": doc.reconciliation_id,
            "message": frappe._("VAT Reconciliation created successfully")
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "VAT Reconciliation Creation Error")
        frappe.throw(str(e))


@frappe.whitelist()
def upload_vat_file(reconciliation_id, file_type):
    """
    Handle CSV file upload for VAT reconciliation
    file_type: 'system', 'itax', or 'tims'
    """
    try:
        # Get uploaded file
        file = frappe.request.files.get('file')
        
        if not file:
            frappe.throw(frappe._("No file uploaded"))

        # Validate file type
        if file_type not in ["system", "itax", "tims"]:
            frappe.throw(frappe._("Invalid file type. Must be 'system', 'itax', or 'tims'"))

        # Get reconciliation document
        doc = frappe.get_doc("VAT Reconciliation", reconciliation_id)

        # Save file
        file_path = save_uploaded_file(file, reconciliation_id, file_type)

        # Validate CSV headers
        validation_result = validate_csv_headers(file_path, file_type)

        if not validation_result["valid"]:
            # Update status to validation failed
            update_file_status(doc, file_type, "Validation Failed", 0)
            doc.save()
            frappe.db.commit()
            
            return {
                "success": False,
                "message": validation_result["message"],
                "missing_columns": validation_result.get("missing_columns", []),
                "found_columns": validation_result.get("found_columns", [])
            }

        # Count rows
        df = pd.read_csv(file_path)
        row_count = len(df)

        # Update document with file reference
        file_field = f"{file_type}_data_file"
        setattr(doc, file_field, file_path)
        update_file_status(doc, file_type, "Validated", row_count)
        
        doc.save()
        frappe.db.commit()

        return {
            "success": True,
            "message": frappe._("File uploaded and validated successfully"),
            "file_path": file_path,
            "row_count": row_count,
            "column_mapping": validation_result.get("column_mapping", {})
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "VAT File Upload Error")
        frappe.throw(str(e))


def save_uploaded_file(file, reconciliation_id, file_type):
    """
    Save uploaded file to files directory
    """
    filename = f"vat_{file_type}_{reconciliation_id}_{file.filename}"
    file_path = frappe.get_site_path('private', 'files', filename)

    # Ensure directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Save file
    file.save(file_path)

    return file_path


def update_file_status(doc, file_type, status, row_count):
    """
    Update file status fields on the document
    """
    status_field = f"{file_type}_data_status"
    rows_field = f"{file_type}_data_rows"
    
    setattr(doc, status_field, status)
    setattr(doc, rows_field, row_count)


def validate_csv_headers(file_path, file_type):
    """
    Validate CSV headers match expected format
    Returns validation result with column mapping
    """
    try:
        df = pd.read_csv(file_path, nrows=1)
        found_columns = [col.lower().strip() for col in df.columns]
        
        expected = EXPECTED_HEADERS[file_type]
        column_mapping = {}
        missing_columns = []

        for required_col in expected["required"]:
            # Check direct match or aliases
            aliases = expected["aliases"].get(required_col, [required_col])
            matched = False
            
            for alias in aliases:
                alias_lower = alias.lower()
                if alias_lower in found_columns:
                    # Find original column name (with original case)
                    original_df = pd.read_csv(file_path, nrows=1)
                    for orig_col in original_df.columns:
                        if orig_col.lower().strip() == alias_lower:
                            column_mapping[required_col] = orig_col
                            matched = True
                            break
                    break
            
            if not matched:
                missing_columns.append(required_col)

        # Check optional columns
        for optional_col in expected.get("optional", []):
            aliases = expected["aliases"].get(optional_col, [optional_col])
            for alias in aliases:
                alias_lower = alias.lower()
                if alias_lower in found_columns:
                    original_df = pd.read_csv(file_path, nrows=1)
                    for orig_col in original_df.columns:
                        if orig_col.lower().strip() == alias_lower:
                            column_mapping[optional_col] = orig_col
                            break
                    break

        if missing_columns:
            return {
                "valid": False,
                "message": frappe._("Missing required columns: {0}").format(", ".join(missing_columns)),
                "missing_columns": missing_columns,
                "found_columns": list(df.columns)
            }

        return {
            "valid": True,
            "message": frappe._("All required columns found"),
            "column_mapping": column_mapping,
            "found_columns": list(df.columns)
        }

    except Exception as e:
        return {
            "valid": False,
            "message": frappe._("Error reading CSV file: {0}").format(str(e)),
            "found_columns": []
        }


@frappe.whitelist()
def parse_csv_data(reconciliation_id, file_type):
    """
    Parse CSV file and store data in child tables
    """
    try:
        doc = frappe.get_doc("VAT Reconciliation", reconciliation_id)
        
        file_field = f"{file_type}_data_file"
        file_path = getattr(doc, file_field)
        
        if not file_path:
            frappe.throw(frappe._("No file uploaded for {0}").format(file_type))

        # Get column mapping
        validation = validate_csv_headers(file_path, file_type)
        if not validation["valid"]:
            frappe.throw(validation["message"])
        
        column_mapping = validation["column_mapping"]
        
        # Read CSV - force cu_invoice_number column to be read as string to prevent scientific notation
        cu_inv_col = column_mapping.get("cu_invoice_number")
        dtype_spec = {cu_inv_col: str} if cu_inv_col else {}
        df = pd.read_csv(file_path, dtype=dtype_spec)
        
        # Clear existing data
        child_table_field = get_child_table_field(file_type)
        doc.set(child_table_field, [])
        
        # Parse and add rows - skip rows without valid cu_invoice_number
        skipped_rows = 0
        added_rows = 0
        for idx, row in df.iterrows():
            child_data = parse_row(row, file_type, column_mapping)
            # Only add rows with valid cu_invoice_number
            if child_data and child_data.get("cu_invoice_number"):
                doc.append(child_table_field, child_data)
                added_rows += 1
            else:
                skipped_rows += 1
        
        # Update status
        update_file_status(doc, file_type, "Parsed", added_rows)
        doc.save()
        frappe.db.commit()

        message = frappe._("Parsed {0} rows successfully").format(added_rows)
        if skipped_rows > 0:
            message += frappe._(". Skipped {0} rows without CU Invoice Number").format(skipped_rows)

        return {
            "success": True,
            "message": message,
            "row_count": added_rows,
            "skipped_rows": skipped_rows
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "CSV Parse Error")
        frappe.throw(str(e))


def get_child_table_field(file_type):
    """Get child table field name for file type"""
    mapping = {
        "system": "system_data",
        "itax": "itax_data",
        "tims": "tims_data"
    }
    return mapping.get(file_type)


def parse_row(row, file_type, column_mapping):
    """Parse a single row based on file type"""
    try:
        posting_date_col = column_mapping.get("posting_date")
        cu_inv_col = column_mapping.get("cu_invoice_number")
        
        # Parse posting date
        posting_date = None
        if posting_date_col and pd.notna(row.get(posting_date_col)):
            try:
                posting_date = pd.to_datetime(row[posting_date_col]).date()
            except:
                posting_date = None
        
        # Normalize CU Invoice Number - strip leading zeros for consistent matching
        cu_invoice_number = ""
        if pd.notna(row.get(cu_inv_col)):
            cu_invoice_number = str(row.get(cu_inv_col, "")).lstrip("0") or "0"

        if file_type == "system":
            amount_col = column_mapping.get("net_amount")
            inv_col = column_mapping.get("invoice_number")
            
            # Parse amount - handle comma-formatted numbers
            amount_value = 0
            if amount_col and pd.notna(row.get(amount_col)):
                amt_str = str(row.get(amount_col, 0)).replace(",", "")
                try:
                    amount_value = float(amt_str)
                except:
                    amount_value = 0
            
            return {
                "posting_date": posting_date,
                "invoice_number": str(row.get(inv_col, "")) if inv_col and pd.notna(row.get(inv_col)) else "",
                "cu_invoice_number": cu_invoice_number,
                "net_amount": amount_value
            }
        
        elif file_type == "itax":
            # iTax CSV has "amount" column which maps to net_amount in DocType
            amount_col = column_mapping.get("amount")
            
            # Parse amount - handle comma-formatted numbers
            amount_value = 0
            if amount_col and pd.notna(row.get(amount_col)):
                amt_str = str(row.get(amount_col, 0)).replace(",", "")
                try:
                    amount_value = float(amt_str)
                except:
                    amount_value = 0
            
            return {
                "posting_date": posting_date,
                "cu_invoice_number": cu_invoice_number,
                "net_amount": amount_value  # Store in net_amount field (DocType field name)
            }
        
        elif file_type == "tims":
            amount_col = column_mapping.get("amount")
            
            # Parse amount - handle comma-formatted numbers
            amount_value = 0
            if amount_col and pd.notna(row.get(amount_col)):
                amt_str = str(row.get(amount_col, 0)).replace(",", "")
                try:
                    amount_value = float(amt_str)
                except:
                    amount_value = 0
            
            return {
                "posting_date": posting_date,
                "cu_invoice_number": cu_invoice_number,
                "amount": amount_value
            }
    
    except Exception as e:
        frappe.log_error(f"Error parsing row: {str(e)}")
        return {}


@frappe.whitelist()
def run_reconciliation(reconciliation_id, use_background=False):
    """
    Run VAT reconciliation between two data sources
    Uses bank reconciliation logic with ±0.01 tolerance
    """
    try:
        doc = frappe.get_doc("VAT Reconciliation", reconciliation_id)
        
        # Check if files are ready
        required_files = doc.get_required_files()
        for file_field, file_name in required_files.items():
            status_field = file_field.replace("_file", "_status")
            if getattr(doc, status_field) not in ["Validated", "Parsed"]:
                frappe.throw(frappe._("{0} must be validated before reconciliation").format(file_name))

        # Parse files if not already parsed
        for file_field in required_files.keys():
            status_field = file_field.replace("_file", "_status")
            if getattr(doc, status_field) == "Validated":
                file_type = file_field.replace("_data_file", "")
                parse_csv_data(reconciliation_id, file_type)

        # Reload document after parsing
        doc.reload()

        # Check total rows to determine if background job needed
        total_rows = get_total_rows(doc)
        
        if total_rows > LARGE_FILE_THRESHOLD or use_background:
            # Queue background job
            doc.is_background_job = 1
            doc.status = "Processing"
            doc.started_at = frappe.utils.now()
            doc.save()
            frappe.db.commit()

            frappe.enqueue(
                "mkaguzi.api.vat_reconciliation.execute_reconciliation",
                queue="long",
                timeout=3600,
                reconciliation_id=reconciliation_id
            )

            return {
                "success": True,
                "background": True,
                "message": frappe._("Reconciliation queued as background job. You will be notified when complete."),
                "job_id": doc.name
            }
        else:
            # Execute immediately
            result = execute_reconciliation(reconciliation_id)
            return result

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "VAT Reconciliation Error")
        frappe.throw(str(e))


def get_total_rows(doc):
    """Get total rows across all parsed data tables"""
    total = 0
    if doc.system_data:
        total += len(doc.system_data)
    if doc.itax_data:
        total += len(doc.itax_data)
    if doc.tims_data:
        total += len(doc.tims_data)
    return total


def execute_reconciliation(reconciliation_id):
    """
    Execute the actual reconciliation logic
    Uses CU Invoice Number as matching key with ±0.01 amount tolerance
    """
    try:
        doc = frappe.get_doc("VAT Reconciliation", reconciliation_id)
        doc.status = "Processing"
        doc.started_at = frappe.utils.now()
        doc.save()
        frappe.db.commit()

        # Get source data based on reconciliation type
        source_a, source_b, source_a_name, source_b_name = get_source_data(doc)

        # Clear existing reconciliation items from database
        frappe.db.delete("VAT Reconciliation Item", {"parent": doc.name})
        frappe.db.commit()
        doc.reconciliation_items = []

        # Build lookup dictionaries
        source_a_dict = build_lookup_dict(source_a, source_a_name)
        source_b_dict = build_lookup_dict(source_b, source_b_name)

        # Track processed items
        processed_b_keys = set()

        # Reconcile: iterate through source A and match with source B
        matched = 0
        unmatched_a = 0
        amount_discrepancies = 0
        total_source_a_amount = 0
        total_source_b_amount = 0
        total_variance = 0  # Sum of absolute variances for all discrepancies
        
        # For large datasets, insert items directly to database in batches
        # to avoid max_allowed_packet issues
        reconciliation_items = []
        BATCH_SIZE = 500
        item_idx = 0
        
        def flush_items():
            nonlocal reconciliation_items, item_idx
            if not reconciliation_items:
                return
            for item_data in reconciliation_items:
                item_data["parent"] = doc.name
                item_data["parenttype"] = "VAT Reconciliation"
                item_data["parentfield"] = "reconciliation_items"
                item_data["doctype"] = "VAT Reconciliation Item"
                item_data["idx"] = item_data.get("idx", 0)
                frappe.get_doc(item_data).db_insert()
            frappe.db.commit()
            reconciliation_items = []

        for cu_inv, items_a in source_a_dict.items():
            for item_a in items_a:
                amount_a = get_amount(item_a, source_a_name)
                total_source_a_amount += amount_a
                
                if cu_inv in source_b_dict:
                    # Found matching CU Invoice Number in source B
                    items_b = source_b_dict[cu_inv]
                    
                    # Find best match (closest amount)
                    best_match = find_best_match(amount_a, items_b, source_b_name, processed_b_keys, cu_inv)
                    
                    if best_match:
                        item_b, amount_b, variance = best_match
                        # Note: amount_b is already added when processing source B totals later
                        
                        # Check if amounts match within tolerance
                        if abs(variance) <= AMOUNT_TOLERANCE:
                            match_status = "Matched"
                            matched += 1
                        else:
                            match_status = "Amount Discrepancy"
                            amount_discrepancies += 1
                            total_variance += abs(variance)

                        # Add reconciliation item
                        item_idx += 1
                        reconciliation_items.append({
                            "idx": item_idx,
                            "cu_invoice_number": cu_inv,
                            "invoice_number": item_a.get("invoice_number", ""),
                            "posting_date_a": item_a.get("posting_date"),
                            "posting_date_b": item_b.get("posting_date"),
                            "amount_a": amount_a,
                            "amount_b": amount_b,
                            "variance": variance,
                            "variance_percentage": (abs(variance) / amount_a * 100) if amount_a else 0,
                            "match_status": match_status
                        })
                    else:
                        # No unprocessed match found in B
                        unmatched_a += 1
                        total_variance += abs(amount_a)  # Add absolute value to total variance
                        
                        item_idx += 1
                        reconciliation_items.append({
                            "idx": item_idx,
                            "cu_invoice_number": cu_inv,
                            "invoice_number": item_a.get("invoice_number", ""),
                            "posting_date_a": item_a.get("posting_date"),
                            "amount_a": amount_a,
                            "amount_b": 0,
                            "variance": amount_a,  # Positive variance (only in A)
                            "variance_percentage": 100,
                            "match_status": "Unmatched Source A"
                        })
                else:
                    # No match in source B
                    unmatched_a += 1
                    total_variance += abs(amount_a)  # Add absolute value to total variance
                    
                    item_idx += 1
                    reconciliation_items.append({
                        "idx": item_idx,
                        "cu_invoice_number": cu_inv,
                        "invoice_number": item_a.get("invoice_number", ""),
                        "posting_date_a": item_a.get("posting_date"),
                        "amount_a": amount_a,
                        "amount_b": 0,
                        "variance": amount_a,  # Positive variance (only in A)
                        "variance_percentage": 100,
                        "match_status": "Unmatched Source A"
                    })
                
                # Flush batch if needed
                if len(reconciliation_items) >= BATCH_SIZE:
                    flush_items()

        # Find unmatched items in source B and calculate total_source_b_amount
        unmatched_b = 0
        for cu_inv, items_b in source_b_dict.items():
            for idx, item_b in enumerate(items_b):
                item_key = f"{cu_inv}_{idx}"
                amount_b = get_amount(item_b, source_b_name)
                
                if item_key not in processed_b_keys:
                    # This item was not matched
                    total_source_b_amount += amount_b
                    unmatched_b += 1
                    total_variance += abs(amount_b)  # Add absolute value to total variance
                    
                    item_idx += 1
                    reconciliation_items.append({
                        "idx": item_idx,
                        "cu_invoice_number": cu_inv,
                        "posting_date_b": item_b.get("posting_date"),
                        "amount_a": 0,
                        "amount_b": amount_b,
                        "variance": -amount_b,  # Negative variance (only in B)
                        "variance_percentage": 100,
                        "match_status": "Unmatched Source B"
                    })
                    
                    # Flush batch if needed
                    if len(reconciliation_items) >= BATCH_SIZE:
                        flush_items()
                else:
                    # This item was matched - add to source B total
                    total_source_b_amount += amount_b
        
        # Flush remaining items
        flush_items()

        # Update summary
        doc.total_source_a_records = sum(len(items) for items in source_a_dict.values())
        doc.total_source_b_records = sum(len(items) for items in source_b_dict.values())
        doc.total_source_a_amount = total_source_a_amount
        doc.total_source_b_amount = total_source_b_amount
        doc.total_matched = matched
        doc.total_unmatched_source_a = unmatched_a
        doc.total_unmatched_source_b = unmatched_b
        doc.total_amount_discrepancies = amount_discrepancies
        doc.total_variance_amount = total_variance

        # Calculate match percentage based on records that found their CU Invoice counterpart
        # Matched = perfect match, Amount Discrepancy = found but different amount
        total_compared = matched + amount_discrepancies + unmatched_a + unmatched_b
        records_matched = matched + amount_discrepancies  # Records where CU Invoice was found in both sources
        doc.match_percentage = (matched / total_compared * 100) if total_compared > 0 else 0

        # Determine reconciliation status
        if doc.match_percentage == 100:
            doc.reconciliation_status = "Fully Reconciled"
        elif doc.match_percentage > 0:
            doc.reconciliation_status = "Partially Reconciled"
        else:
            doc.reconciliation_status = "Not Reconciled"

        doc.status = "Completed"
        doc.completed_at = frappe.utils.now()
        doc.progress_percent = 100
        
        # Use db_update to update only parent fields without triggering child table save
        # This avoids max_allowed_packet issues for large reconciliations
        doc.db_update()
        frappe.db.commit()

        # Compare with previous month if requested
        if doc.compare_with_previous:
            doc.compare_with_previous_month()

        # Send email notification for background jobs
        if doc.is_background_job:
            send_completion_notification(doc)

        return {
            "success": True,
            "message": frappe._("Reconciliation completed successfully"),
            "summary": {
                "total_matched": matched,
                "total_unmatched_source_a": unmatched_a,
                "total_unmatched_source_b": unmatched_b,
                "total_amount_discrepancies": amount_discrepancies,
                "total_variance_amount": total_variance,
                "match_percentage": doc.match_percentage,
                "reconciliation_status": doc.reconciliation_status
            }
        }

    except Exception as e:
        # Update status to failed
        doc = frappe.get_doc("VAT Reconciliation", reconciliation_id)
        doc.status = "Failed"
        doc.error_message = str(e)
        doc.save()
        frappe.db.commit()
        
        frappe.log_error(frappe.get_traceback(), "VAT Reconciliation Execution Error")
        raise


def get_source_data(doc):
    """Get source data based on reconciliation type"""
    if doc.reconciliation_type == "System vs iTax":
        return doc.system_data, doc.itax_data, "system", "itax"
    elif doc.reconciliation_type == "System vs TIMs":
        return doc.system_data, doc.tims_data, "system", "tims"
    elif doc.reconciliation_type == "iTax vs TIMs":
        return doc.itax_data, doc.tims_data, "itax", "tims"
    return [], [], "", ""


def build_lookup_dict(data, source_name):
    """Build lookup dictionary keyed by CU Invoice Number"""
    lookup = {}
    for item in data:
        cu_inv = item.cu_invoice_number
        if cu_inv:
            if cu_inv not in lookup:
                lookup[cu_inv] = []
            lookup[cu_inv].append(item.as_dict())
    return lookup


def get_amount(item, source_name):
    """Get amount field based on source type"""
    if source_name == "system":
        # System data uses net_amount
        return float(item.get("net_amount", 0))
    elif source_name == "itax":
        # iTax data also uses net_amount (field name in DocType)
        return float(item.get("net_amount", 0))
    elif source_name == "tims":
        # TIMs data uses amount
        return float(item.get("amount", 0))
    return 0


def find_best_match(amount_a, items_b, source_b_name, processed_keys, cu_inv):
    """Find best matching item from source B based on amount"""
    best_match = None
    best_variance = float('inf')
    best_idx = None

    for idx, item_b in enumerate(items_b):
        item_key = f"{cu_inv}_{idx}"
        if item_key in processed_keys:
            continue
        
        amount_b = get_amount(item_b, source_b_name)
        variance = amount_a - amount_b

        if abs(variance) < abs(best_variance):
            best_variance = variance
            best_match = item_b
            best_idx = idx

    if best_match is not None:
        processed_keys.add(f"{cu_inv}_{best_idx}")
        return (best_match, get_amount(best_match, source_b_name), best_variance)
    
    return None


def send_completion_notification(doc):
    """Send email notification when background job completes"""
    try:
        user = frappe.get_doc("User", doc.prepared_by)
        
        subject = frappe._("VAT Reconciliation {0} Completed").format(doc.reconciliation_id)
        message = frappe._("""
            <p>Your VAT Reconciliation has been completed.</p>
            <p><strong>Details:</strong></p>
            <ul>
                <li>Reconciliation ID: {0}</li>
                <li>Type: {1}</li>
                <li>Month: {2}</li>
                <li>Status: {3}</li>
                <li>Match Percentage: {4}%</li>
                <li>Total Variance: {5}</li>
            </ul>
            <p><a href="/app/vat-reconciliation/{6}">View Reconciliation</a></p>
        """).format(
            doc.reconciliation_id,
            doc.reconciliation_type,
            doc.reconciliation_month,
            doc.reconciliation_status,
            round(doc.match_percentage, 2),
            frappe.format_value(doc.total_variance_amount, {"fieldtype": "Currency"}),
            doc.name
        )

        frappe.sendmail(
            recipients=[user.email],
            subject=subject,
            message=message
        )
    except Exception as e:
        frappe.log_error(f"Failed to send notification: {str(e)}")


@frappe.whitelist(allow_guest=False)
def get_reconciliation_results(reconciliation_id, filter_status=None, page=1, page_size=100):
    """
    Get reconciliation results with optional filtering and database-level pagination
    Optimized for large datasets by using database queries instead of loading all items into memory
    """
    try:
        page = int(page)
        page_size = int(page_size)
        offset = (page - 1) * page_size

        # Build filters
        filters = {"parent": reconciliation_id}
        if filter_status:
            filters["match_status"] = filter_status

        # Get total count for pagination
        total_count = frappe.db.count("VAT Reconciliation Item", filters=filters)

        # Get paginated results from database
        items = frappe.get_all(
            "VAT Reconciliation Item",
            filters=filters,
            fields=[
                "name", "cu_invoice_number", "invoice_number",
                "posting_date_a", "posting_date_b",
                "amount_a", "amount_b", "variance", "variance_percentage",
                "match_status", "idx"
            ],
            order_by="idx",
            limit=page_size,
            start=offset
        )

        # Get summary fields directly from database without loading child tables
        # This avoids loading 8000+ child records into memory
        summary = frappe.db.get_value(
            "VAT Reconciliation",
            reconciliation_id,
            [
                "total_matched", "total_unmatched_source_a", "total_unmatched_source_b",
                "total_amount_discrepancies", "total_variance_amount",
                "match_percentage", "reconciliation_status"
            ],
            as_dict=True
        )

        return {
            "items": items,
            "total_count": total_count,
            "page": page,
            "page_size": page_size,
            "total_pages": (total_count + page_size - 1) // page_size,
            "summary": summary or {}
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Reconciliation Results Error")
        frappe.throw(str(e))


@frappe.whitelist(allow_guest=False)
def get_reconciliation_summary(reconciliation_id):
    """
    Get reconciliation summary data using lightweight database query.
    This avoids loading child tables (which can have 8000+ records) by using
    frappe.db.get_value() instead of frappe.get_doc().
    
    Optimized for fast initial page load in the frontend.
    """
    try:
        # Get all parent document fields without loading child tables
        summary = frappe.db.get_value(
            "VAT Reconciliation",
            reconciliation_id,
            [
                "name", "reconciliation_id", "reconciliation_type", "status",
                "reconciliation_month", "fiscal_year", "reconciliation_date",
                "total_source_a_records", "total_source_b_records",
                "total_source_a_amount", "total_source_b_amount",
                "total_matched", "total_unmatched_source_a", "total_unmatched_source_b",
                "total_amount_discrepancies", "total_variance_amount", "match_percentage",
                "reconciliation_status", "prepared_by", "is_background_job",
                "progress_percent", "completed_at", "creation", "modified",
                "system_data_file", "itax_data_file", "tims_data_file",
                "system_data_status", "itax_data_status", "tims_data_status",
                "system_data_rows", "itax_data_rows", "tims_data_rows",
                "started_at", "error_message"
            ],
            as_dict=True
        )
        
        if not summary:
            frappe.throw(_("VAT Reconciliation {0} not found").format(reconciliation_id))
        
        return {
            "success": True,
            "data": summary
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Reconciliation Summary Error")
        frappe.throw(str(e))


@frappe.whitelist()
def get_reconciliation_progress(reconciliation_id):
    """
    Get progress of background reconciliation job
    """
    doc = frappe.get_doc("VAT Reconciliation", reconciliation_id)
    
    return {
        "status": doc.status,
        "progress_percent": doc.progress_percent,
        "is_background_job": doc.is_background_job,
        "started_at": doc.started_at,
        "completed_at": doc.completed_at,
        "error_message": doc.error_message
    }


@frappe.whitelist()
def export_reconciliation_report(reconciliation_id, format="excel"):
    """
    Export reconciliation report to Excel or PDF
    Optimized to avoid loading all child records - passes only reconciliation_id
    """
    try:
        # Check if user has read permission on the document
        if not frappe.has_permission("VAT Reconciliation", "read", reconciliation_id):
            frappe.throw(frappe._("You do not have permission to export this reconciliation"), frappe.PermissionError)
        
        # Don't load full document - pass reconciliation_id directly to avoid loading 8000+ child records
        if format == "excel":
            return export_to_excel(reconciliation_id)
        elif format == "pdf":
            return export_to_pdf(reconciliation_id)
        else:
            frappe.throw(frappe._("Unsupported export format: {0}").format(format))

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Export Report Error")
        frappe.throw(str(e))


def export_to_excel(reconciliation_id):
    """
    Export reconciliation to Excel with structured multi-sheet report:
    - Sheet 1: Summary (reconciliation overview and statistics)
    - Sheet 2: Missing in A (Unmatched Source A - records in B but not in A)
    - Sheet 3: Missing in B (Unmatched Source B - records in A but not in B)
    - Sheet 4: Amount Mismatch (records with variance between sources)
    
    Optimized to use database queries instead of loading all child records.
    """
    try:
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        # Get summary data from database (lightweight query)
        summary = frappe.db.get_value(
            "VAT Reconciliation",
            reconciliation_id,
            [
                "name", "reconciliation_id", "reconciliation_type", "status",
                "reconciliation_month", "fiscal_year", "reconciliation_date",
                "total_source_a_records", "total_source_b_records",
                "total_source_a_amount", "total_source_b_amount",
                "total_matched", "total_unmatched_source_a", "total_unmatched_source_b",
                "total_amount_discrepancies", "total_variance_amount", "match_percentage",
                "reconciliation_status", "completed_at"
            ],
            as_dict=True
        )
        
        if not summary:
            frappe.throw(_("VAT Reconciliation {0} not found").format(reconciliation_id))
        
        # Define column structure for item sheets
        item_columns = [
            "cu_invoice_number", "invoice_number", "posting_date_a", "posting_date_b",
            "amount_a", "amount_b", "variance", "variance_percentage"
        ]
        item_display_names = {
            "cu_invoice_number": "CU Invoice Number",
            "invoice_number": "Invoice Number",
            "posting_date_a": "Date (Source A)",
            "posting_date_b": "Date (Source B)",
            "amount_a": "Amount (Source A)",
            "amount_b": "Amount (Source B)",
            "variance": "Variance",
            "variance_percentage": "Variance %"
        }
        
        # Fetch items by status using optimized database queries
        # Convert frappe._dict to regular dict for pandas compatibility
        def get_items_by_status(status):
            items = frappe.get_all(
                "VAT Reconciliation Item",
                filters={"parent": reconciliation_id, "match_status": status},
                fields=item_columns,
                order_by="cu_invoice_number"
            )
            # Convert to regular dicts for pandas compatibility
            return [dict(item) for item in items]
        
        missing_in_a = get_items_by_status("Unmatched Source A")
        missing_in_b = get_items_by_status("Unmatched Source B")
        amount_mismatch = get_items_by_status("Amount Discrepancy")
        
        # Create Excel workbook
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # ===== SHEET 1: SUMMARY =====
            summary_data = {
                "Field": [
                    "Report Title",
                    "",
                    "RECONCILIATION DETAILS",
                    "Reconciliation ID",
                    "Reconciliation Type",
                    "Period",
                    "Reconciliation Date",
                    "Status",
                    "Completed At",
                    "",
                    "RECORD COUNTS",
                    "Total Source A Records",
                    "Total Source B Records",
                    "Total Matched",
                    "Missing in A (Unmatched Source A)",
                    "Missing in B (Unmatched Source B)",
                    "Amount Discrepancies",
                    "Match Percentage",
                    "",
                    "FINANCIAL SUMMARY",
                    "Total Source A Amount",
                    "Total Source B Amount",
                    "Total Variance Amount"
                ],
                "Value": [
                    "VAT RECONCILIATION REPORT",
                    "",
                    "",
                    summary.reconciliation_id or summary.name,
                    summary.reconciliation_type,
                    f"{summary.reconciliation_month} {summary.fiscal_year}",
                    str(summary.reconciliation_date) if summary.reconciliation_date else "-",
                    summary.reconciliation_status or summary.status,
                    str(summary.completed_at) if summary.completed_at else "-",
                    "",
                    "",
                    summary.total_source_a_records or 0,
                    summary.total_source_b_records or 0,
                    summary.total_matched or 0,
                    summary.total_unmatched_source_a or 0,
                    summary.total_unmatched_source_b or 0,
                    summary.total_amount_discrepancies or 0,
                    f"{(summary.match_percentage or 0):.2f}%",
                    "",
                    "",
                    f"KES {(summary.total_source_a_amount or 0):,.2f}",
                    f"KES {(summary.total_source_b_amount or 0):,.2f}",
                    f"KES {(summary.total_variance_amount or 0):,.2f}"
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='1. Summary', index=False)
            
            # ===== SHEET 2: MISSING IN A (Unmatched Source A) =====
            if missing_in_a:
                missing_a_df = pd.DataFrame(missing_in_a)
                missing_a_df.columns = [item_display_names.get(c, c) for c in missing_a_df.columns]
                # Add subtotals
                subtotal_row = pd.DataFrame([{
                    "CU Invoice Number": f"SUBTOTAL ({len(missing_in_a)} records)",
                    "Invoice Number": "",
                    "Date (Source A)": "",
                    "Date (Source B)": "",
                    "Amount (Source A)": sum(item.get("amount_a") or 0 for item in missing_in_a),
                    "Amount (Source B)": sum(item.get("amount_b") or 0 for item in missing_in_a),
                    "Variance": sum(item.get("variance") or 0 for item in missing_in_a),
                    "Variance %": ""
                }])
                missing_a_df = pd.concat([missing_a_df, subtotal_row], ignore_index=True)
            else:
                missing_a_df = pd.DataFrame([{"Message": "No records missing in Source A"}])
            missing_a_df.to_excel(writer, sheet_name='2. Missing in A', index=False)
            
            # ===== SHEET 3: MISSING IN B (Unmatched Source B) =====
            if missing_in_b:
                missing_b_df = pd.DataFrame(missing_in_b)
                missing_b_df.columns = [item_display_names.get(c, c) for c in missing_b_df.columns]
                # Add subtotals
                subtotal_row = pd.DataFrame([{
                    "CU Invoice Number": f"SUBTOTAL ({len(missing_in_b)} records)",
                    "Invoice Number": "",
                    "Date (Source A)": "",
                    "Date (Source B)": "",
                    "Amount (Source A)": sum(item.get("amount_a") or 0 for item in missing_in_b),
                    "Amount (Source B)": sum(item.get("amount_b") or 0 for item in missing_in_b),
                    "Variance": sum(item.get("variance") or 0 for item in missing_in_b),
                    "Variance %": ""
                }])
                missing_b_df = pd.concat([missing_b_df, subtotal_row], ignore_index=True)
            else:
                missing_b_df = pd.DataFrame([{"Message": "No records missing in Source B"}])
            missing_b_df.to_excel(writer, sheet_name='3. Missing in B', index=False)
            
            # ===== SHEET 4: AMOUNT MISMATCH =====
            if amount_mismatch:
                mismatch_df = pd.DataFrame(amount_mismatch)
                mismatch_df.columns = [item_display_names.get(c, c) for c in mismatch_df.columns]
                # Add subtotals
                subtotal_row = pd.DataFrame([{
                    "CU Invoice Number": f"SUBTOTAL ({len(amount_mismatch)} records)",
                    "Invoice Number": "",
                    "Date (Source A)": "",
                    "Date (Source B)": "",
                    "Amount (Source A)": sum(item.get("amount_a") or 0 for item in amount_mismatch),
                    "Amount (Source B)": sum(item.get("amount_b") or 0 for item in amount_mismatch),
                    "Variance": sum(item.get("variance") or 0 for item in amount_mismatch),
                    "Variance %": ""
                }])
                mismatch_df = pd.concat([mismatch_df, subtotal_row], ignore_index=True)
            else:
                mismatch_df = pd.DataFrame([{"Message": "No amount mismatches found"}])
            mismatch_df.to_excel(writer, sheet_name='4. Amount Mismatch', index=False)
            
            # Apply styling to all sheets
            workbook = writer.book
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_font = Font(bold=True, size=14)
            section_font = Font(bold=True, size=11, color="2F5496")
            subtotal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            subtotal_font = Font(bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Style Summary sheet
            ws_summary = workbook['1. Summary']
            ws_summary.column_dimensions['A'].width = 35
            ws_summary.column_dimensions['B'].width = 40
            
            # Style title
            ws_summary['B1'].font = title_font
            
            # Style section headers
            for row in [3, 11, 20]:
                ws_summary.cell(row=row, column=1).font = section_font
            
            # Style item sheets
            for sheet_name in ['2. Missing in A', '3. Missing in B', '4. Amount Mismatch']:
                ws = workbook[sheet_name]
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Style header row
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                
                # Style data rows and subtotal
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    for cell in row:
                        cell.border = border
                        # Check if this is the subtotal row (last row)
                        if row_idx == ws.max_row and ws.cell(row=row_idx, column=1).value and "SUBTOTAL" in str(ws.cell(row=row_idx, column=1).value):
                            cell.fill = subtotal_fill
                            cell.font = subtotal_font

        output.seek(0)
        
        # Save as file using Frappe's file handling for proper permissions
        filename = f"VAT_Reconciliation_{summary.reconciliation_id or reconciliation_id}_{summary.reconciliation_month}_{summary.fiscal_year}.xlsx"
        
        # Create a File document for proper permission handling
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": filename,
            "is_private": 1,
            "content": output.getvalue(),
            "attached_to_doctype": "VAT Reconciliation",
            "attached_to_name": reconciliation_id
        })
        file_doc.save(ignore_permissions=True)
        
        return {
            "success": True,
            "file_path": file_doc.file_url,
            "filename": filename,
            "download_url": file_doc.file_url
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Excel Export Error")
        frappe.throw(str(e))


def export_to_pdf(reconciliation_id):
    """Export reconciliation to PDF - optimized to avoid loading full document"""
    try:
        # Get summary data from database (lightweight query)
        doc = frappe.db.get_value(
            "VAT Reconciliation",
            reconciliation_id,
            [
                "name", "reconciliation_id", "reconciliation_type", "status",
                "reconciliation_month", "fiscal_year", "reconciliation_date",
                "total_source_a_records", "total_source_b_records",
                "total_source_a_amount", "total_source_b_amount",
                "total_matched", "total_unmatched_source_a", "total_unmatched_source_b",
                "total_amount_discrepancies", "total_variance_amount", "match_percentage",
                "reconciliation_status"
            ],
            as_dict=True
        )
        
        if not doc:
            frappe.throw(_("VAT Reconciliation {0} not found").format(reconciliation_id))
        
        # Get reconciliation items by status (only non-matched items for the report)
        items = frappe.get_all(
            "VAT Reconciliation Item",
            filters={"parent": reconciliation_id, "match_status": ["!=", "Matched"]},
            fields=["*"],
            order_by="match_status, cu_invoice_number"
        )
        
        # Build HTML report
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>VAT Reconciliation Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 12px; margin: 20px; }}
                h1 {{ color: #333; font-size: 18px; margin-bottom: 10px; }}
                h2 {{ color: #555; font-size: 14px; margin-top: 20px; margin-bottom: 10px; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f4f4f4; font-weight: bold; }}
                .summary {{ background-color: #f9f9f9; padding: 15px; margin-bottom: 20px; }}
                .summary-item {{ display: inline-block; margin-right: 30px; }}
                .matched {{ color: green; }}
                .unmatched {{ color: orange; }}
                .discrepancy {{ color: red; }}
                .text-right {{ text-align: right; }}
                .header-info {{ margin-bottom: 20px; }}
                .header-info p {{ margin: 5px 0; }}
            </style>
        </head>
        <body>
            <h1>VAT Reconciliation Report</h1>
            
            <div class="header-info">
                <p><strong>Reconciliation ID:</strong> {doc.reconciliation_id or doc.name}</p>
                <p><strong>Type:</strong> {doc.reconciliation_type}</p>
                <p><strong>Period:</strong> {doc.reconciliation_month} {doc.fiscal_year}</p>
                <p><strong>Status:</strong> {doc.reconciliation_status or doc.status}</p>
                <p><strong>Generated:</strong> {frappe.utils.now()}</p>
            </div>
            
            <div class="summary">
                <h2>Summary</h2>
                <div class="summary-item"><strong>Source A Records:</strong> {doc.total_source_a_records or 0}</div>
                <div class="summary-item"><strong>Source B Records:</strong> {doc.total_source_b_records or 0}</div>
                <div class="summary-item"><strong>Matched:</strong> <span class="matched">{doc.total_matched or 0}</span></div>
                <div class="summary-item"><strong>Unmatched A:</strong> <span class="unmatched">{doc.total_unmatched_source_a or 0}</span></div>
                <div class="summary-item"><strong>Unmatched B:</strong> <span class="unmatched">{doc.total_unmatched_source_b or 0}</span></div>
                <div class="summary-item"><strong>Discrepancies:</strong> <span class="discrepancy">{doc.total_amount_discrepancies or 0}</span></div>
                <br><br>
                <div class="summary-item"><strong>Source A Amount:</strong> {frappe.utils.fmt_money(doc.total_source_a_amount or 0, currency='KES')}</div>
                <div class="summary-item"><strong>Source B Amount:</strong> {frappe.utils.fmt_money(doc.total_source_b_amount or 0, currency='KES')}</div>
                <div class="summary-item"><strong>Total Variance:</strong> <span class="discrepancy">{frappe.utils.fmt_money(doc.total_variance_amount or 0, currency='KES')}</span></div>
                <div class="summary-item"><strong>Match Rate:</strong> {(doc.match_percentage or 0):.1f}%</div>
            </div>
            
            <h2>Reconciliation Details</h2>
            <table>
                <thead>
                    <tr>
                        <th>CU Invoice No</th>
                        <th>Invoice No</th>
                        <th>Date A</th>
                        <th>Date B</th>
                        <th class="text-right">Amount A</th>
                        <th class="text-right">Amount B</th>
                        <th class="text-right">Variance</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for item in items:
            status_class = ""
            if item.match_status == "Matched":
                status_class = "matched"
            elif "Unmatched" in (item.match_status or ""):
                status_class = "unmatched"
            else:
                status_class = "discrepancy"
            
            html += f"""
                    <tr>
                        <td>{item.cu_invoice_number or ''}</td>
                        <td>{item.invoice_number or ''}</td>
                        <td>{item.posting_date_a or ''}</td>
                        <td>{item.posting_date_b or ''}</td>
                        <td class="text-right">{frappe.utils.fmt_money(item.amount_a or 0, currency='KES')}</td>
                        <td class="text-right">{frappe.utils.fmt_money(item.amount_b or 0, currency='KES')}</td>
                        <td class="text-right">{frappe.utils.fmt_money(item.variance or 0, currency='KES')}</td>
                        <td class="{status_class}">{item.match_status or ''}</td>
                    </tr>
            """
        
        html += """
                </tbody>
            </table>
        </body>
        </html>
        """

        # Try to generate PDF, fall back to HTML if wkhtmltopdf not available
        try:
            from frappe.utils.pdf import get_pdf
            pdf_content = get_pdf(html)
            
            filename = f"VAT_Reconciliation_{doc.reconciliation_id or reconciliation_id}_{doc.reconciliation_month}_{doc.fiscal_year}.pdf"
            
            # Create a File document for proper permission handling
            file_doc = frappe.get_doc({
                "doctype": "File",
                "file_name": filename,
                "is_private": 1,
                "content": pdf_content,
                "attached_to_doctype": "VAT Reconciliation",
                "attached_to_name": reconciliation_id
            })
            file_doc.save(ignore_permissions=True)

            return {
                "success": True,
                "file_path": file_doc.file_url,
                "filename": filename,
                "download_url": file_doc.file_url
            }
        except OSError as pdf_error:
            # wkhtmltopdf not available, save as HTML instead
            frappe.log_error(str(pdf_error), "PDF Generation - wkhtmltopdf not found, using HTML")
            
            filename = f"VAT_Reconciliation_{doc.reconciliation_id or reconciliation_id}_{doc.reconciliation_month}_{doc.fiscal_year}.html"
            
            # Create a File document for proper permission handling
            file_doc = frappe.get_doc({
                "doctype": "File",
                "file_name": filename,
                "is_private": 1,
                "content": html.encode('utf-8'),
                "attached_to_doctype": "VAT Reconciliation",
                "attached_to_name": reconciliation_id
            })
            file_doc.save(ignore_permissions=True)

            return {
                "success": True,
                "file_path": file_doc.file_url,
                "filename": filename,
                "download_url": file_doc.file_url,
                "format": "html",
                "message": "PDF generation requires wkhtmltopdf. Report exported as HTML instead."
            }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "PDF Export Error")
        frappe.throw(str(e))


@frappe.whitelist()
def compare_with_previous_month(reconciliation_id):
    """
    Compare current reconciliation with previous month
    """
    try:
        doc = frappe.get_doc("VAT Reconciliation", reconciliation_id)
        result = doc.compare_with_previous_month()
        
        if result:
            return {
                "success": True,
                "comparison": result,
                "previous_reconciliation": doc.previous_reconciliation,
                "variance_trend": doc.variance_trend
            }
        else:
            return {
                "success": False,
                "message": frappe._("No previous reconciliation found for comparison")
            }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Historical Comparison Error")
        frappe.throw(str(e))


@frappe.whitelist()
def get_recurring_discrepancies(reconciliation_id, months=6):
    """
    Find CU Invoice Numbers that appear as discrepancies across multiple months
    """
    try:
        doc = frappe.get_doc("VAT Reconciliation", reconciliation_id)
        
        # Get discrepancy CU Invoice Numbers from current reconciliation
        current_discrepancies = set()
        for item in doc.reconciliation_items:
            if item.match_status in ["Unmatched Source A", "Unmatched Source B", "Amount Discrepancy"]:
                current_discrepancies.add(item.cu_invoice_number)

        # Find previous reconciliations
        previous_recs = frappe.get_all(
            "VAT Reconciliation",
            filters={
                "reconciliation_type": doc.reconciliation_type,
                "status": ["in", ["Completed", "Reviewed", "Approved"]],
                "name": ["!=", doc.name]
            },
            fields=["name", "reconciliation_month", "fiscal_year"],
            order_by="creation desc",
            limit=months
        )

        recurring = {}
        for cu_inv in current_discrepancies:
            occurrences = []
            
            for prev_rec in previous_recs:
                prev_doc = frappe.get_doc("VAT Reconciliation", prev_rec.name)
                for item in prev_doc.reconciliation_items:
                    if item.cu_invoice_number == cu_inv and item.match_status != "Matched":
                        occurrences.append({
                            "month": prev_rec.reconciliation_month,
                            "year": prev_rec.fiscal_year,
                            "status": item.match_status,
                            "variance": item.variance
                        })
                        break

            if occurrences:
                recurring[cu_inv] = occurrences

        return {
            "success": True,
            "recurring_discrepancies": recurring,
            "count": len(recurring)
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Recurring Discrepancies Error")
        frappe.throw(str(e))


@frappe.whitelist()
def delete_reconciliation(reconciliation_id):
    """Delete a VAT Reconciliation document"""
    try:
        doc = frappe.get_doc("VAT Reconciliation", reconciliation_id)
        
        if doc.status in ["Approved", "Reviewed"]:
            frappe.throw(frappe._("Cannot delete approved or reviewed reconciliations"))

        doc.delete()
        frappe.db.commit()

        return {
            "success": True,
            "message": frappe._("Reconciliation deleted successfully")
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Delete Reconciliation Error")
        frappe.throw(str(e))
